from .forms_event_creation import ActivityForm, RegistrationFieldForm, CoordinatorForm, WebsiteSetupForm
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

# Step 2: Activities
@login_required
def step2_activities(request):
    activities = request.session.get('activities', [])
    if request.method == 'POST':
        form = ActivityForm(request.POST)
        if form.is_valid():
            activities.append(form.cleaned_data)
            request.session['activities'] = activities
            if 'add_another' in request.POST:
                return redirect('step2_activities')
            else:
                return redirect('step3_registration_form')
    else:
        form = ActivityForm()
    return render(request, 'events/step2_activities.html', {'form': form, 'activities': activities})

# Step 3: Registration Form Builder
@login_required
def step3_registration_form(request):
    fields = request.session.get('registration_fields', [])
    if request.method == 'POST':
        form = RegistrationFieldForm(request.POST)
        if form.is_valid():
            fields.append(form.cleaned_data)
            request.session['registration_fields'] = fields
            if 'add_another' in request.POST:
                return redirect('step3_registration_form')
            else:
                return redirect('step4_coordinators')
    else:
        form = RegistrationFieldForm()
    return render(request, 'events/step3_registration_form.html', {'form': form, 'fields': fields})

# Step 4: Coordinators
@login_required
def step4_coordinators(request):
    coordinators = request.session.get('coordinators', [])
    if request.method == 'POST':
        form = CoordinatorForm(request.POST)
        if form.is_valid():
            coordinators.append(form.cleaned_data)
            request.session['coordinators'] = coordinators
            if 'add_another' in request.POST:
                return redirect('step4_coordinators')
            else:
                return redirect('step5_website_setup')
    else:
        form = CoordinatorForm()
    return render(request, 'events/step4_coordinators.html', {'form': form, 'coordinators': coordinators})

# Step 5: Website Setup
@login_required
def step5_website_setup(request):
    if request.method == 'POST':
        form = WebsiteSetupForm(request.POST)
        if form.is_valid():
            request.session['website_setup'] = form.cleaned_data
            # Save all data to DB here
            return redirect('event_creation_complete')
    else:
        form = WebsiteSetupForm()
    return render(request, 'events/step5_website_setup.html', {'form': form})

@login_required
def event_creation_complete(request):
    return render(request, 'events/event_creation_complete.html')

from django.contrib.auth.decorators import login_required
from .forms_event_creation import ActivityForm, RegistrationFieldForm, CoordinatorForm, WebsiteSetupForm

# Step 2: Activities
@login_required
def step2_activities(request):
    activities = request.session.get('activities', [])
    if request.method == 'POST':
        form = ActivityForm(request.POST)
        if form.is_valid():
            activities.append(form.cleaned_data)
            request.session['activities'] = activities
            if 'add_another' in request.POST:
                return redirect('step2_activities')
            else:
                return redirect('step3_registration_form')
    else:
        form = ActivityForm()
    return render(request, 'events/step2_activities.html', {'form': form, 'activities': activities})

# Step 3: Registration Form Builder
@login_required
def step3_registration_form(request):
    fields = request.session.get('registration_fields', [])
    if request.method == 'POST':
        form = RegistrationFieldForm(request.POST)
        if form.is_valid():
            fields.append(form.cleaned_data)
            request.session['registration_fields'] = fields
            if 'add_another' in request.POST:
                return redirect('step3_registration_form')
            else:
                return redirect('step4_coordinators')
    else:
        form = RegistrationFieldForm()
    return render(request, 'events/step3_registration_form.html', {'form': form, 'fields': fields})

# Step 4: Coordinators
@login_required
def step4_coordinators(request):
    coordinators = request.session.get('coordinators', [])
    if request.method == 'POST':
        form = CoordinatorForm(request.POST)
        if form.is_valid():
            coordinators.append(form.cleaned_data)
            request.session['coordinators'] = coordinators
            if 'add_another' in request.POST:
                return redirect('step4_coordinators')
            else:
                return redirect('step5_website_setup')
    else:
        form = CoordinatorForm()
    return render(request, 'events/step4_coordinators.html', {'form': form, 'coordinators': coordinators})

# Step 5: Website Setup
@login_required
def step5_website_setup(request):
    if request.method == 'POST':
        form = WebsiteSetupForm(request.POST)
        if form.is_valid():
            request.session['website_setup'] = form.cleaned_data
            # Save all data to DB here
            return redirect('event_creation_complete')
    else:
        form = WebsiteSetupForm()
    return render(request, 'events/step5_website_setup.html', {'form': form})

@login_required
def event_creation_complete(request):
    return render(request, 'events/event_creation_complete.html')
from .models import (
    Activity,
    Event,
    EventCoordinator,
    EventCoordinatorInvite,
    Profile,
    ActivityRegistrationFormField,
    ActivityRegistrationFormResponse,
    ActivityCoordinator,
)
from django.shortcuts import render, redirect, get_object_or_404
# Activity Registration Form View
from participant.models import ActivityRegistration
from django.contrib.auth.decorators import login_required
from .auth_decorators import (
    require_role,
    require_organizer,
    require_coordinator,
    require_participant,
    get_user_dashboard_redirect,
    get_user_role,
)

@login_required
def activity_registration_form(request, activity_id):
    activity = get_object_or_404(Activity, id=activity_id)
    if not _is_participant(request.user):
        messages.error(request, "Only participants can register for activities.")
        return redirect("events:dashboard")

    # If you have custom fields for activities, parse them here
    fields = []
    if hasattr(activity, "registration_fields"):
        fields = activity.registration_fields.all().order_by("order", "id")

        if request.method == "POST":
            form_data = {}
            missing = []
            for field in fields:
                value = (request.POST.get(field.label) or "").strip()
                if field.required and not value:
                    missing.append(field.label)
                form_data[field.label] = value

            payment_qr = request.FILES.get('payment_qr')
            if not payment_qr:
                messages.error(request, "Payment QR screenshot is required.")
                return render(request, "events/activity_registration.html", {"activity": activity, "fields": fields})

            if missing:
                messages.error(request, f"Please provide required fields: {', '.join(missing)}.")
            else:
                # Capacity check: respect activity.max_participants when set
                if activity.max_participants:
                    current_count = ActivityRegistration.objects.filter(activity=activity).count()
                    if current_count >= activity.max_participants:
                        messages.error(request, "Registration Closed: this activity has reached the maximum number of participants.")
                        return render(request, "events/activity_registration.html", {"activity": activity, "fields": fields})

                # prevent duplicate registrations
                if ActivityRegistration.objects.filter(participant=request.user, activity=activity).exists():
                    messages.error(request, "You have already registered for this activity.")
                    return redirect("events:dashboard")

                # Capacity check: respect activity.max_participants when set
                if activity.max_participants:
                    current_count = ActivityRegistration.objects.filter(activity=activity).count()
                    if current_count >= activity.max_participants:
                        messages.error(request, "Registration Closed: this activity has reached the maximum number of participants.")
                        return render(request, "events/activity_registration.html", {"activity": activity, "fields": fields})

                registration = ActivityRegistration.objects.create(
                    participant=request.user,
                    activity=activity,
                    qr_token=uuid.uuid4(),
                )
                # attach form data if model supports it
                if hasattr(registration, 'form_data'):
                    try:
                        registration.form_data = form_data
                        registration.save(update_fields=['form_data'])
                    except Exception:
                        pass

                # store payment proof
                payment_check, _ = PaymentCheck.objects.get_or_create(registration=registration)
                payment_check.payment_proof = payment_qr
                payment_check.save()
                messages.success(request, "Registration + payment proof submitted.")
                return redirect("events:dashboard")

    return render(
        request,
        "events/activity_registration.html",
        {
            "activity": activity,
            "fields": fields,
        },
    )
import io
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET

@login_required
@require_GET
def generate_gate_pass(request, registration_id):
    from participant.models import EventRegistration
    registration = get_object_or_404(EventRegistration, id=registration_id, participant=request.user)
    event = registration.event
    # Generate QR code (contains immutable token + user + event)
    qr_token = uuid.uuid5(uuid.NAMESPACE_URL, f"event-pass-{registration.id}-{request.user.id}")
    qr_data = f"EVTPASS:{qr_token}|USER:{request.user.id}|EVENT:{event.id}"
    qr_img = qrcode.make(qr_data)
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    # Generate PDF
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 18)
    c.drawString(30 * mm, height - 40 * mm, f"Gate Pass for {event.event}")
    c.setFont("Helvetica", 12)
    c.drawString(30 * mm, height - 55 * mm, f"Participant: {request.user.get_full_name() or request.user.username}")
    c.drawString(30 * mm, height - 65 * mm, f"Event Date: {event.date_of_event}")
    c.drawString(30 * mm, height - 75 * mm, f"Venue: {event.venue}")
    # Draw QR code
    c.drawInlineImage(qr_buffer, 30 * mm, height - 120 * mm, 50 * mm, 50 * mm)
    c.setFont("Helvetica", 10)
    c.drawString(30 * mm, height - 130 * mm, f"Registration ID: {registration.id}")
    c.drawString(30 * mm, height - 140 * mm, f"Token: {qr_token}")
    c.showPage()
    c.save()
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename=gate_pass_{registration.id}.pdf'
    return response
import base64
import json
import os
import uuid
import urllib.error
import urllib.request
from io import BytesIO
import re

from django.contrib.auth import authenticate, get_user_model, login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.db import IntegrityError, models
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, HttpResponseNotAllowed, JsonResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.contrib.sessions.models import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    Activity,
    Event,
    EventCoordinator,
    EventCoordinatorInvite,
    Profile,
    ActivityRegistrationFormField,
    ActivityRegistrationFormResponse,
    ActivityCoordinator,
)
from .forms import EventForm
from participant.models import (
    ActivityRegistration,
    AttendanceRecord,
    Certificate,
    EventRegistration,
    LeaderboardEntry,
    TeamRegistration,
)
from payment.models import PaymentCheck
from notification.services import (
    send_activity_registration_email,
    send_event_registration_email,
)


def _can_manage_event(user, event):
    return user.is_superuser or user.is_staff or event.user_id == user.id


def _is_organizer(user):
    if not user.is_authenticated:
        return False
    return Profile.objects.filter(user=user, role="organizer").exists()


def _is_coordinator(user):
    if not user.is_authenticated:
        return False
    return Profile.objects.filter(user=user, role="coordinator").exists()


def _get_coordinator_role(user):
    if not user.is_authenticated:
        return None
    role = (
        Profile.objects.filter(user=user, role="coordinator")
        .values_list("coordinator_role", flat=True)
        .first()
    )
    if role is None and _is_coordinator(user):
        return "activity"
    return role


def _is_head_coordinator(user):
    return _get_coordinator_role(user) == "head"


def _is_event_coordinator(user):
    return _get_coordinator_role(user) == "event"


def _is_activity_coordinator(user):
    return _get_coordinator_role(user) == "activity"


def _is_participant(user):
    if not user.is_authenticated:
        return False
    return Profile.objects.filter(user=user, role="participant").exists()


def _can_check_payment_for_registration(user, registration):
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if _is_organizer(user):
        return True
    if _is_head_coordinator(user):
        return True
    if _is_event_coordinator(user):
        return EventCoordinator.objects.filter(
            event=registration.activity.event,
            user=user,
        ).exists()
    if not _is_coordinator(user):
        return False
    return ActivityCoordinator.objects.filter(
        activity=registration.activity,
        user=user,
    ).exists()


def _is_event_coordinator_for_event(user, event):
    if not user.is_authenticated:
        return False
    if _is_head_coordinator(user):
        return True
    if not _is_event_coordinator(user):
        return False
    return EventCoordinator.objects.filter(event=event, user=user).exists()


def _is_activity_coordinator_for_activity(user, activity):
    if not user.is_authenticated:
        return False
    if _is_head_coordinator(user):
        return True
    if _is_event_coordinator(user):
        return EventCoordinator.objects.filter(event=activity.event, user=user).exists()
    if not _is_activity_coordinator(user):
        return False
    return ActivityCoordinator.objects.filter(activity=activity, user=user).exists()


def _is_activity_coordinator_for_event(user, event):
    if not user.is_authenticated:
        return False
    if _is_head_coordinator(user):
        return True
    if _is_event_coordinator(user):
        return EventCoordinator.objects.filter(event=event, user=user).exists()
    if not _is_activity_coordinator(user):
        return False
    return ActivityCoordinator.objects.filter(activity__event=event, user=user).exists()


def _can_manage_event_activities(user, event):
    return (
        user.is_superuser
        or user.is_staff
        or (event.user_id == user.id)
        or _is_event_coordinator_for_event(user, event)
    )


def _can_view_event_participants(user, event):
    return (
        user.is_superuser
        or user.is_staff
        or event.user_id == user.id
        or _is_event_coordinator_for_event(user, event)
    )


def _can_manage_activity(user, activity):
    if user.is_superuser or user.is_staff:
        return True
    if _is_organizer(user) and activity.event.user_id == user.id:
        return True
    return _is_activity_coordinator_for_activity(user, activity)


def _can_view_event_dashboard(user, event):
    if not user.is_authenticated:
        return False
    return _can_manage_event_activities(user, event) or _is_activity_coordinator_for_event(
        user,
        event,
    )


def _authenticate_for_role(request, username, password, expected_role):
    # Try authenticating by username first, then fall back to email lookup.
    user = authenticate(request, username=username, password=password)
    if user is None:
        # Allow users to login with their email address as well
        UserModel = get_user_model()
        candidate = UserModel.objects.filter(email__iexact=(username or "").strip()).first()
        if candidate:
            user = authenticate(request, username=candidate.username, password=password)

    if user is None:
        return None, "Invalid username or password. Please try again."

    if user.is_superuser:
        return user, None

    profile_role = (
        Profile.objects.filter(user=user).values_list("role", flat=True).first()
    )
    if profile_role != expected_role:
        return (
            None,
            f"This account is not registered as {expected_role.title()}.",
        )

    return user, None


def _get_openai_key():
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is not set.")
    return api_key


def _openai_post(endpoint, payload):
    api_key = _get_openai_key()
    url = f"https://api.openai.com/v1/{endpoint}"
    data = json.dumps(payload).encode("utf-8")
    request_obj = urllib.request.Request(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request_obj, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"OpenAI API error: {exc.code} {body}") from exc


def _extract_response_text(data):
    output = data.get("output", []) if isinstance(data, dict) else []
    parts = []
    for item in output:
        for content in item.get("content", []) or []:
            if content.get("type") in {"output_text", "text"}:
                parts.append(content.get("text", ""))
    return "\n".join([part for part in parts if part]).strip()


def _parse_json_object(text):
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(text[start : end + 1])
        raise


def _parse_registration_fields(text):
    fields = []
    if not text:
        return fields

    for line in text.splitlines():
        label = line.strip()
        if not label:
            continue
        required = label.endswith("*")
        if required:
            label = label[:-1].strip()
        if not label:
            continue
        fields.append(
            {
                "key": f"field_{len(fields) + 1}",
                "label": label,
                "required": required,
            }
        )
    return fields


def _normalize_registration_fields(raw_text):
    if not raw_text:
        return "", None

    cleaned = []
    seen = set()
    for line in raw_text.splitlines():
        label = line.strip()
        if not label:
            continue
        if label.endswith("*"):
            label = label[:-1].strip() + "*"
        if not label or label == "*":
            return None, "Registration field labels cannot be empty."
        lower_label = label.lower()
        if lower_label in seen:
            return None, "Duplicate registration field labels are not allowed."
        seen.add(lower_label)
        cleaned.append(label)

    return "\n".join(cleaned), None


def _extract_event_payload(request):
    data = {
        "event": request.POST.get("event") or request.GET.get("event"),
        "activity": request.POST.get("activity") or request.GET.get("activity"),
        "description": request.POST.get("description") or request.GET.get("description"),
        "tagline": request.POST.get("tagline") or request.GET.get("tagline"),
        "schedule": request.POST.get("schedule") or request.GET.get("schedule"),
        "activities_overview": request.POST.get("activities_overview")
        or request.GET.get("activities_overview"),
        "rules": request.POST.get("rules") or request.GET.get("rules"),
        "date_of_event": request.POST.get("date_of_event")
        or request.GET.get("date_of_event"),
        "time_of_event": request.POST.get("time_of_event")
        or request.GET.get("time_of_event"),
        "venue": request.POST.get("venue") or request.GET.get("venue"),
        "category": request.POST.get("category") or request.GET.get("category"),
        "registration": request.POST.get("registration_fees")
        or request.GET.get("registration_fees")
        or request.POST.get("registration")
        or request.GET.get("registration"),
        "announcement": request.POST.get("announcement")
        or request.GET.get("announcement"),
        "contact_info": request.POST.get("contact_info")
        or request.GET.get("contact_info"),
        "image_url": request.POST.get("image_url") or request.GET.get("image_url"),
        "template_choice": request.POST.get("template_choice")
        or request.GET.get("template_choice"),
    }

    # For JSON API calls where request.POST is empty.
    if (
        request.content_type
        and "application/json" in request.content_type
        and not any(data.values())
    ):
        try:
            import json

            body = json.loads(request.body.decode("utf-8") or "{}")
            data = {
                "event": body.get("event"),
                "activity": body.get("activity"),
                "description": body.get("description"),
                "tagline": body.get("tagline"),
                "schedule": body.get("schedule"),
                "activities_overview": body.get("activities_overview"),
                "rules": body.get("rules"),
                "date_of_event": body.get("date_of_event"),
                "time_of_event": body.get("time_of_event"),
                "venue": body.get("venue"),
                "category": body.get("category"),
                "registration": body.get("registration_fees")
                or body.get("registration"),
                "announcement": body.get("announcement"),
                "contact_info": body.get("contact_info"),
                "image_url": body.get("image_url"),
                "template_choice": body.get("template_choice"),
            }
        except (ValueError, UnicodeDecodeError):
            return None

    return data


def _send_activity_invite_email(request, invite):
    accept_path = reverse("events:accept_activity_invite", args=[str(invite.token)])
    accept_url = request.build_absolute_uri(accept_path)
    signup_url = request.build_absolute_uri(
        f"{reverse('events:signup')}?role=coordinator&next={accept_path}"
    )

    subject = f"Activity invite for {invite.activity.name}"
    message = (
        f"You have been invited to coordinate the activity: {invite.activity.name}.\n"
        f"Event: {invite.activity.event.event}.\n\n"
        f"Accept invite: {accept_url}\n"
        f"If you need an account, sign up here: {signup_url}\n\n"
        "This invite is tied to this email address."
    )

    from_email = settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL
    send_mail(subject, message, from_email, [invite.email], fail_silently=False)


def _send_event_invite_email(request, invite):
    accept_path = reverse("events:accept_event_invite", args=[str(invite.token)])
    accept_url = request.build_absolute_uri(accept_path)
    signup_url = request.build_absolute_uri(
        f"{reverse('events:signup')}?role=coordinator&next={accept_path}"
    )

    subject = f"Event coordinator invite for {invite.event.event}"
    message = (
        f"You have been invited to coordinate the event: {invite.event.event}.\n\n"
        f"Accept invite: {accept_url}\n"
        f"If you need an account, sign up here: {signup_url}\n\n"
        "This invite is tied to this email address."
    )

    from_email = settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL
    send_mail(subject, message, from_email, [invite.email], fail_silently=False)


@csrf_exempt
@login_required
def event_list(request):
    if request.method == "GET":
        q = request.GET.get('q', '').strip()
        date_filter = request.GET.get('date', '').strip()
        category = request.GET.get('category', '').strip()
        fee = request.GET.get('fee', '').strip().lower()

        if request.user.is_superuser or request.user.is_staff:
            queryset = Event.objects.select_related("user").all().order_by("id")
        elif _is_organizer(request.user):
            queryset = (
                Event.objects.select_related("user").filter(user=request.user).order_by("id")
            )
        else:
            # Participants and visitors: show all events (read-only)
            queryset = Event.objects.select_related("user").all().order_by("id")

        if q:
            queryset = queryset.filter(event__icontains=q)
        if date_filter:
            queryset = queryset.filter(date_of_event=date_filter)
        if category:
            queryset = queryset.filter(category=category)
        if fee:
            if fee == 'free':
                queryset = queryset.filter(registration__iregex=r'(?i)free')
            elif fee == 'paid':
                queryset = queryset.exclude(registration__iregex=r'(?i)free')

        # For traditional browser requests, render the event list template
        return render(
            request,
            "events/event_list.html",
            {
                "events": queryset,
                "can_create_event": _is_organizer(request.user),
                "search_q": q,
                "filter_date": date_filter,
                "filter_category": category,
                "filter_fee": fee,
            },
        )

    if request.method == "POST":
        if not _is_organizer(request.user):
            return JsonResponse(
                {"error": "Only organizers can create events."},
                status=403,
            )

        payload = _extract_event_payload(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload."}, status=400)

        missing = [key for key, value in payload.items() if not value]
        missing = [
            key
            for key in missing
            if key
            not in {
                "announcement",
                "tagline",
                "schedule",
                "activities_overview",
                "rules",
            }
        ]
        if missing:
            return JsonResponse(
                {"error": f"Missing required fields: {', '.join(missing)}."},
                status=400,
            )

        created = Event.objects.create(user=request.user, **payload)

        return JsonResponse(
            {
                "id": created.id,
                "event": created.event,
                "activity": created.activity,
                "date_of_event": created.date_of_event,
                "time_of_event": created.time_of_event,
                "registration": created.registration,
                "registration_fees": created.registration,
                "announcement": created.announcement,
            },
            status=201,
        )

    return HttpResponseNotAllowed(["GET", "POST"])


@csrf_exempt #removed login required temp
def event_detail(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    # Determine if the currently authenticated user is already registered
    from participant.models import EventRegistration

    is_registered = EventRegistration.objects.filter(participant=request.user, event=event).exists()

    return render(
        request,
        "events/event_detail.html",
        {
            "event": event,
            "can_manage": _can_manage_event(request.user, event),
            "is_registered": is_registered,
        },
    )

    if request.method in {"PUT", "PATCH"}:
        payload = _extract_event_payload(request)
        if payload is None:
            return JsonResponse({"error": "Invalid JSON payload."}, status=400)

        for key, value in payload.items():
            if value:
                setattr(event, key, value)
        event.save()

        return JsonResponse(
            {
                "id": event.id,
                "event": event.event,
                "activity": event.activity,
                "description": event.description,
                "tagline": event.tagline,
                "schedule": event.schedule,
                "activities_overview": event.activities_overview,
                "rules": event.rules,
                "date_of_event": event.date_of_event,
                "time_of_event": event.time_of_event,
                "registration": event.registration,
                "announcement": event.announcement,
                "venue": event.venue,
                "category": event.category,
                "contact_info": event.contact_info,
                "image_url": event.image_url,
                "template_choice": event.template_choice,
            },
            status=200,
        )

    if request.method == "DELETE":
        event.delete()
        return JsonResponse({"message": "Event deleted successfully."}, status=200)

    return HttpResponseNotAllowed(["GET", "PUT", "PATCH", "DELETE"])


@login_required
def ai_generate_event_content(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    if not _is_organizer(request.user):
        return JsonResponse({"error": "Only organizers can generate content."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except (ValueError, UnicodeDecodeError):
        payload = {}

    title = (payload.get("event") or payload.get("title") or "").strip()
    category = (payload.get("category") or payload.get("type") or "").strip()
    venue = (payload.get("venue") or "").strip()
    date_of_event = (payload.get("date_of_event") or "").strip()

    if not title or not category:
        return JsonResponse(
            {"error": "Event title and category are required."},
            status=400,
        )

    prompt = (
        "You are an assistant for an event organizer. Generate event content as JSON.\n"
        "Return ONLY valid JSON with this shape:\n"
        "{\n"
        "  \"tagline\": string,\n"
        "  \"description\": string,\n"
        "  \"schedule\": [string],\n"
        "  \"activities\": [string],\n"
        "  \"rules\": [string]\n"
        "}\n"
        "Keep it concise and upbeat. Avoid placeholders.\n"
        f"Event title: {title}\n"
        f"Event category/type: {category}\n"
        f"Venue (if known): {venue or 'Unknown'}\n"
        f"Date (if known): {date_of_event or 'Unknown'}\n"
    )

    try:
        response = _openai_post(
            "responses",
            {
                "model": "gpt-4.1-mini",
                "input": [
                    {
                        "role": "user",
                        "content": [{"type": "input_text", "text": prompt}],
                    }
                ],
                "temperature": 0.7,
            },
        )
        text = _extract_response_text(response)
        if not text:
            raise RuntimeError("Empty response from OpenAI.")
        data = _parse_json_object(text)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=502)

    return JsonResponse(
        {
            "tagline": data.get("tagline", ""),
            "description": data.get("description", ""),
            "schedule": data.get("schedule", []),
            "activities": data.get("activities", []),
            "rules": data.get("rules", []),
        }
    )


@login_required
def ai_generate_event_banner(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    if not _is_organizer(request.user):
        return JsonResponse({"error": "Only organizers can generate banners."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except (ValueError, UnicodeDecodeError):
        payload = {}

    title = (payload.get("event") or payload.get("title") or "").strip()
    category = (payload.get("category") or payload.get("type") or "").strip()

    if not title or not category:
        return JsonResponse(
            {"error": "Event title and category are required."},
            status=400,
        )

    prompt = (
        "Create a premium event banner image with a clean, modern look. "
        "No text, no logos, no watermarks. "
        f"Theme: {category}. Event: {title}. "
        "Use a cinematic atmosphere with soft lighting and rich detail."
    )

    try:
        response = _openai_post(
            "images",
            {
                "model": "gpt-image-1",
                "prompt": prompt,
                "size": "1024x1024",
                "response_format": "b64_json",
            },
        )
        data_list = response.get("data", []) if isinstance(response, dict) else []
        if not data_list:
            raise RuntimeError("Empty image response from OpenAI.")
        image_b64 = data_list[0].get("b64_json")
        if not image_b64:
            raise RuntimeError("Image data missing from OpenAI response.")

        image_bytes = base64.b64decode(image_b64)
        filename = f"events/ai/banner_{uuid.uuid4().hex}.png"
        saved_path = default_storage.save(filename, ContentFile(image_bytes))
        image_url = default_storage.url(saved_path)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=502)

    return JsonResponse({"image_url": image_url})


def role_based_user_page(request):
    organizer_exists = Profile.objects.filter(role="organizer").exists()
    if request.user.is_authenticated and request.user.is_superuser:
        role = "Super Admin (Platform Owner)"
        message = "You can manage organizers, templates, and all events across the platform."
    elif request.user.is_authenticated:
        profile_role = (
            Profile.objects.filter(user=request.user)
            .values_list("role", flat=True)
            .first()
        )
        if profile_role == "coordinator":
            role = "Coordinator"
            message = "You can coordinate schedules, activities, and registrations."
        elif profile_role == "organizer":
            role = "Organizer"
            message = "You can manage events, announcements, and platform operations."
        elif profile_role == "participant":
            role = "Participant"
            message = "You can browse, register, and track your event participation."
        elif request.user.is_staff:
            role = "Staff"
            message = "You can manage event operations and participant updates."
        else:
            role = "User"
            message = "You can browse, register, and track your event participation."
    else:
        role = "Guest"
        message = "Sign up to create events and manage your registrations."

    return render(
        request,
        "index.html",
        {
            "role": role,
            "message": message,
            "organizer_exists": organizer_exists,
        },
    )


def signup(request):
    user_model = get_user_model()
    valid_roles = {key for key, _ in Profile.ROLE_CHOICES}
    valid_coordinator_roles = {key for key, _ in Profile.COORDINATOR_ROLE_CHOICES}
    selected_role = request.GET.get("role") or request.session.get("selected_role")
    
    # If no role selected, or invalid role, send user to the dedicated chooser
    if not selected_role or selected_role not in valid_roles:
        return redirect("accounts:role_selection")

    # Coordinators are NOT allowed to sign up directly
    if selected_role == "coordinator":
        messages.error(request, "Coordinators cannot sign up directly. You must be invited by an Organizer.")
        return redirect("accounts:role_selection")

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        email = (request.POST.get("email") or "").strip()
        password1 = request.POST.get("password1") or ""
        password2 = request.POST.get("password2") or ""
        role = request.POST.get("role") or ""
        coordinator_role = request.POST.get("coordinator_role") or ""
        errors = []

        if not username:
            errors.append("Username is required.")
        if not email:
            errors.append("Email is required.")
        if not password1:
            errors.append("Password is required.")
        if password1 != password2:
            errors.append("Passwords do not match.")
        if role not in valid_roles:
            errors.append("Please select a valid role.")
        if role == "coordinator":
            errors.append("Coordinators cannot sign up directly.")
        if username and user_model.objects.filter(username=username).exists():
            errors.append("Username is already taken.")
        if email and user_model.objects.filter(email=email).exists():
            errors.append("Email is already registered.")

        if not errors:
            user = user_model.objects.create_user(
                username=username,
                email=email,
                password=password1,
            )
            Profile.objects.create(
                user=user,
                role=role,
                coordinator_role=coordinator_role if role == "coordinator" else None,
            )
            request.session["selected_role"] = role
            login(request, user)
            # Redirect to role-specific dashboard
            dashboard_redirect = get_user_dashboard_redirect(user)
            return redirect(dashboard_redirect)

        return render(
            request,
            "registration/signup.html",
            {
                "errors": errors,
                "username": username,
                "email": email,
                "selected_role": role,
                "selected_coordinator_role": coordinator_role,
                "roles": Profile.ROLE_CHOICES,
                "coordinator_roles": Profile.COORDINATOR_ROLE_CHOICES,
            },
        )

    return render(
        request,
        "registration/signup.html",
        {
            "selected_role": selected_role,
            "roles": Profile.ROLE_CHOICES,
            "coordinator_roles": Profile.COORDINATOR_ROLE_CHOICES,
        },
    )


@login_required
def invite_coordinator(request):
    if not _is_organizer(request.user):
        messages.error(request, "Only organizers can invite coordinators.")
        return redirect("events:dashboard")

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        email = (request.POST.get("email") or "").strip().lower()
        coordinator_type = (request.POST.get("coordinator_type") or "").strip()
        event_id = request.POST.get("event")

        if not name or not email or not coordinator_type:
            messages.error(request, "Please provide name, email, and coordinator type.")
            return redirect("events:organizer_dashboard")
            
        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Please enter a valid email address.")
            return redirect("events:organizer_dashboard")

        event = None
        if event_id:
            event = Event.objects.filter(id=event_id, user=request.user).first()
            if not event:
                messages.error(request, "Invalid event selected or you do not have permission for this event.")
                return redirect("events:organizer_dashboard")

        # Create the invite record
        if event:
            invite, created = EventCoordinatorInvite.objects.get_or_create(
                event=event,
                email=email,
                status=EventCoordinatorInvite.STATUS_PENDING,
                defaults={"invited_by": request.user},
            )
            
            # Trigger email properly with safeguards as requested
            try:
                _send_event_invite_email(request, invite)
                if created:
                    messages.success(request, f"Invitation sent to {email} as {coordinator_type}.")
                else:
                    messages.info(request, f"Invitation already existed for {email}; email resent.")
            except Exception as e:
                # Log error or print (non-blocking as requested)
                print(f"Failed to send invite email to {email}: {str(e)}")
                messages.warning(request, f"Invitation recorded for {name}, but email failed to send. Check system logs.")
        else:
             # General coordinating invites without specific event might be allowed in future,
             # but for now EventCoordinatorInvite model requires an event link.
             messages.error(request, "A specific event must be selected to invite a coordinator.")

        return redirect("events:organizer_dashboard")
        
    return redirect("events:organizer_dashboard")



def unified_login(request):
    """Unified login page for all user roles."""
    selected_role = request.GET.get("role") or request.session.get("selected_role") or "coordinator"
    signup_selected_role = request.GET.get("signup_role") or selected_role or "organizer"
    context = {
        "selected_role": selected_role,
        "signup_selected_role": signup_selected_role,
        "roles": Profile.ROLE_CHOICES,
        "username": "",
        "error_message": None,
    }

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        role = request.POST.get("role") or "coordinator"

        if not username or not password:
            context["error_message"] = "Username and password are required."
            context["username"] = username
            context["selected_role"] = role
        else:
            user, error_message = _authenticate_for_role(
                request,
                username,
                password,
                role,
            )
            if user is not None:
                login(request, user)
                request.session["selected_role"] = role
                # Redirect to role-specific dashboard
                dashboard_redirect = get_user_dashboard_redirect(user)
                return redirect(dashboard_redirect)
            else:
                context["error_message"] = error_message
                context["username"] = username
                context["selected_role"] = role

    return render(request, "registration/unified_login.html", context)


def role_login(request, expected_role, template_name):
    context = {
        "username": "",
        "error_message": None,
    }

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""

        context["username"] = username

        if not username or not password:
            context["error_message"] = "Username and password are required."
            return render(request, template_name, context)

        user, error_message = _authenticate_for_role(
            request,
            username,
            password,
            expected_role,
        )
        if user is None:
            context["error_message"] = error_message
            return render(request, template_name, context)

        login(request, user)
        # Redirect to role-specific dashboard
        dashboard_redirect = get_user_dashboard_redirect(user)
        return redirect(dashboard_redirect)

    return render(request, template_name, context)


@login_required
def manage_events(request):
    can_create_event = _is_organizer(request.user)
    form = EventForm()

    if request.method == "POST":
        if not can_create_event:
            messages.error(request, "Only organizers can create events.")
            return redirect("events:manage_events")

        form = EventForm(request.POST, request.FILES)
        images = request.FILES.getlist("images")
        
        if form.is_valid():
            created_event = form.save(commit=False)
            created_event.user = request.user
            try:
                created_event.save()
            except IntegrityError:
                form.add_error("name", "An event with this name already exists.")
                messages.error(request, "An event with this name already exists.")
            else:
                # Save multiple images if provided
                from .models_eventimage import EventImage
                if images:
                    for img in images:
                        EventImage.objects.create(event=created_event, image=img)
                
                messages.success(request, "Event created successfully.")
                return redirect("event_site", event_id=created_event.id)
        else:
            # If form is invalid, we return the errors to the template.
            # The template will need to be updated to show these.
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').capitalize()}: {error}")

    if request.user.is_superuser or request.user.is_staff:
        events = Event.objects.all().order_by("-id")
    elif _is_organizer(request.user):
        events = Event.objects.filter(user=request.user).order_by("-id")
    else:
        events = Event.objects.none()

    return render(
        request,
        "events/manage_events.html",
        {
            "form": form,
            "can_create_event": can_create_event,
            "category_choices": Event.CATEGORY_CHOICES,
            "template_choices": Event.TEMPLATE_CHOICES,
            "events": events,
        },
    )


@login_required
def edit_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    can_edit = request.user.is_superuser or request.user.is_staff or (
        _is_organizer(request.user) and event.user_id == request.user.id
    )
    if not can_edit:
        messages.error(request, "You do not have permission to edit this event.")
        return redirect("events:manage_events")

    form = EventForm(instance=event)

    if request.method == "POST":
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            try:
                form.save()
            except IntegrityError:
                form.add_error("name", "An event with this name already exists.")
                messages.error(request, "An event with this name already exists.")
            else:
                messages.success(request, "Event updated successfully.")
                return redirect("events:edit_event", event_id=event.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').capitalize()}: {error}")

    return render(
        request,
        "events/edit_event.html",
        {
            "form": form,
            "event": event,
            "category_choices": Event.CATEGORY_CHOICES,
            "template_choices": Event.TEMPLATE_CHOICES,
        },
    )

    return render(
        request,
        "events/edit_event.html",
        {
            "event": event,
            "category_choices": Event.CATEGORY_CHOICES,
            "template_choices": Event.TEMPLATE_CHOICES,
        },
    )


@login_required
def event_registration_form(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    # Only participants may register; if user hasn't chosen a role, send them to role selection
    if not _is_participant(request.user):
        messages.error(request, "Only participants can register for events. Please choose your role.")
        return redirect("accounts:role_selection")

    # Prevent an organizer from registering for their own event
    if _is_organizer(request.user) and event.user_id == request.user.id:
        messages.error(request, "Organizers cannot register for their own event.")
        return redirect("events:event_detail", event_id=event.id)

    fields = _parse_registration_fields(event.registration_form_fields)
    for field in fields:
        field["value"] = ""
    existing_registration = EventRegistration.objects.filter(
        participant=request.user,
        event=event,
    ).first()
    form_data = existing_registration.form_data if existing_registration and existing_registration.form_data else {}
    for field in fields:
        field["value"] = form_data.get(field["label"]) or form_data.get(field["key"]) or ""
    submission_complete = existing_registration is not None
    download_pass_url = (
        reverse("events:generate_gate_pass", args=[existing_registration.id])
        if existing_registration
        else None
    )

    if request.method == "POST":
        form_data = dict(form_data) if isinstance(form_data, dict) else {}
        missing = []

        full_name = (request.POST.get("full_name") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        email = (request.POST.get("email") or request.user.email or "").strip()
        payment_reference = (request.POST.get("payment_reference") or "").strip()

        if not full_name:
            missing.append("Full Name")
        if not phone:
            missing.append("Phone Number")
        if not email:
            missing.append("Email")
        if not payment_reference:
            missing.append("Payment Reference Number")

        def _save_uploaded_file(upload, folder):
            if not upload:
                return ""
            filename = default_storage.save(f"registration_uploads/{folder}/{upload.name}", upload)
            return filename

        id_proof = _save_uploaded_file(request.FILES.get("id_proof"), "id_proof")
        payment_screenshot = _save_uploaded_file(request.FILES.get("payment_screenshot"), "payment")
        photo = _save_uploaded_file(request.FILES.get("photo"), "photos")

        if not id_proof:
            missing.append("ID Proof")
        if not payment_screenshot:
            missing.append("Payment Screenshot")

        form_data.update(
            {
                "full_name": full_name,
                "phone": phone,
                "email": email,
                "payment_reference": payment_reference,
                "id_proof": id_proof,
                "payment_screenshot": payment_screenshot,
                "photo": photo,
            }
        )

        for field in fields:
            value = (request.POST.get(field["key"]) or "").strip()
            if field["required"] and not value:
                missing.append(field["label"])
            form_data[field["label"]] = value
            form_data[field["key"]] = value

        if missing:
            messages.error(
                request,
                f"Please provide required fields: {', '.join(missing)}.",
            )
        else:
            # Do not allow registration if event completed
            if event.status == Event.STATUS_COMPLETED:
                messages.error(request, "Registration is closed for completed events.")
                return redirect("events:event_detail", event_id=event.id)

            registration, created = EventRegistration.objects.get_or_create(
                participant=request.user,
                event=event,
                defaults={"form_data": form_data},
            )
            if not created:
                registration.form_data = form_data
                registration.save(update_fields=["form_data"])
                messages.info(request, "Registration updated.")
            else:
                messages.success(request, "You are registered for this event.")
                send_event_registration_email(request.user, event)
            existing_registration = registration
            submission_complete = True
            download_pass_url = reverse("events:generate_gate_pass", args=[registration.id])

    return render(
        request,
        "events/event_registration.html",
        {
            "event": event,
            "fields": fields,
            "registration": existing_registration,
            "submission_complete": submission_complete,
            "download_pass_url": download_pass_url,
            "registration_data": form_data,
        },
    )


@login_required
def export_event_excel(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    can_export = request.user.is_superuser or request.user.is_staff
    if not can_export:
        can_export = _is_organizer(request.user) and event.user_id == request.user.id
    # Allow activity coordinators to export if they coordinate any activity in this event
    if not can_export:
        is_activity_coordinator = ActivityCoordinator.objects.filter(
            user=request.user, activity__event=event
        ).exists()
        can_export = is_activity_coordinator
    if not can_export:
        return HttpResponse("You do not have permission to export this event.", status=403)

    activity_coordinators = ActivityCoordinator.objects.select_related("user")
    activity_coordinator_map = {}
    for row in activity_coordinators:
        activity_coordinator_map.setdefault(row.activity_id, []).append(row.user)

    activity_registration_counts = {
        row["activity_id"]: row["count"]
        for row in ActivityRegistration.objects.filter(activity__event=event)
        .values("activity_id")
        .annotate(count=models.Count("id"))
    }
    team_registration_counts = {
        row["activity_id"]: row["count"]
        for row in TeamRegistration.objects.filter(activity__event=event)
        .values("activity_id")
        .annotate(count=models.Count("id"))
    }

    wb = Workbook()
    # Add a new sheet for Activity Participants
    ws5 = wb.create_sheet("Activity Participants")
    ws5.append([
        "Activity Name",
        "Participant Username",
        "Participant Email",
        "Registered At",
        "Event Name",
        "Event Date",
        "Start Time",
        "End Time",
        "Registration Fee",
        "Team Event",
        "Team Size",
    ])
    activities = Activity.objects.select_related("event").filter(event=event).order_by("id")
    for activity in activities:
        # Ensure registrations are filtered by both activity and event
        registrations = ActivityRegistration.objects.select_related("participant", "activity").filter(activity=activity, activity__event=event)
        for reg in registrations:
            ws5.append([
                activity.name,
                reg.participant.username,
                reg.participant.email,
                reg.registered_at.isoformat(),
                activity.event.event,
                activity.event.date_of_event.isoformat() if activity.event.date_of_event else "",
                activity.start_time.isoformat() if activity.start_time else "",
                activity.end_time.isoformat() if activity.end_time else "",
                activity.registration_fee,
                "Yes" if activity.is_team_event else "No",
                activity.team_size or "",
            ])
        if not registrations.exists():
            ws5.append([
                activity.name,
                "No participants yet.", "", "", activity.event.event,
                activity.event.date_of_event.isoformat() if activity.event.date_of_event else "",
                activity.start_time.isoformat() if activity.start_time else "",
                activity.end_time.isoformat() if activity.end_time else "",
                activity.registration_fee,
                "Yes" if activity.is_team_event else "No",
                activity.team_size or "",
            ])

    # Fix event registrations export to ensure correct event
    ws = wb.active
    ws.title = "Event Registrations"
    ws.append(["Participant", "Email", "Event", "Registered At", "Form Data"])
    event_registrations = EventRegistration.objects.select_related("participant", "event").filter(event=event)
    if event_registrations.exists():
        for reg in event_registrations:
            ws.append([
                reg.participant.username,
                reg.participant.email,
                reg.event.event,
                reg.registered_at.isoformat(),
                json.dumps(reg.form_data or {}, ensure_ascii=False),
            ])
    else:
        ws.append(["No event registrations yet.", "", "", "", ""])
    ws = wb.active
    ws.title = "Event Registrations"
    ws.append(["Participant", "Email", "Event", "Registered At", "Form Data"])
    event_registrations = EventRegistration.objects.select_related("participant").filter(
        event=event
    )
    if event_registrations.exists():
        for reg in event_registrations:
            ws.append(
                [
                    reg.participant.username,
                    reg.participant.email,
                    event.event,
                    reg.registered_at.isoformat(),
                    json.dumps(reg.form_data or {}, ensure_ascii=False),
                ]
            )
    else:
        ws.append(["No event registrations yet.", "", "", "", ""]) 

    ws2 = wb.create_sheet("Activities")
    ws2.append(
        [
            "Activity",
            "Event",
            "Event Date",
            "Start Time",
            "End Time",
            "Registration Fee",
            "Team Event",
            "Team Size",
            "Coordinators",
            "Registrations",
            "Teams",
        ]
    )
    activities = Activity.objects.select_related("event").filter(event=event).order_by("id")
    if activities.exists():
        for activity in activities:
            coordinators = activity_coordinator_map.get(activity.id, [])
            coordinator_names = ", ".join(user.username for user in coordinators) or "-"
            ws2.append(
                [
                    activity.name,
                    event.event,
                    event.date_of_event.isoformat() if event.date_of_event else "",
                    activity.start_time.isoformat() if activity.start_time else "",
                    activity.end_time.isoformat() if activity.end_time else "",
                    activity.registration_fee,
                    "Yes" if activity.is_team_event else "No",
                    activity.team_size or "",
                    coordinator_names,
                    activity_registration_counts.get(activity.id, 0),
                    team_registration_counts.get(activity.id, 0),
                ]
            )
    else:
        ws2.append(["No activities created yet.", "", "", "", "", "", "", "", "", "", ""]) 

    ws3 = wb.create_sheet("Activity Registrations")
    ws3.append(
        [
            "Participant",
            "Email",
            "Activity",
            "Event",
            "Event Date",
            "Time",
            "Coordinators",
            "Registered At",
            "Payment Status",
        ]
    )
    payment_map = {
        row.registration_id: row
        for row in PaymentCheck.objects.select_related("registration")
    }
    for reg in ActivityRegistration.objects.select_related(
        "participant", "activity", "activity__event"
    ).filter(activity__event=event):
        coordinators = activity_coordinator_map.get(reg.activity_id, [])
        coordinator_names = ", ".join(user.username for user in coordinators) or "-"
        time_range = ""
        if reg.activity.start_time and reg.activity.end_time:
            time_range = f"{reg.activity.start_time} - {reg.activity.end_time}"
        elif reg.activity.start_time:
            time_range = str(reg.activity.start_time)
        elif reg.activity.end_time:
            time_range = str(reg.activity.end_time)
        payment = payment_map.get(reg.id)
        ws3.append(
            [
                reg.participant.username,
                reg.participant.email,
                reg.activity.name,
                reg.activity.event.event,
                reg.activity.event.date_of_event.isoformat()
                if reg.activity.event.date_of_event
                else "",
                time_range,
                coordinator_names,
                reg.registered_at.isoformat(),
                payment.status if payment else "pending",
            ]
        )

    if ws3.max_row == 1:
        ws3.append(["No activity registrations yet.", "", "", "", "", "", "", "", ""]) 

    ws4 = wb.create_sheet("Team Registrations")
    ws4.append(
        [
            "Team Name",
            "Leader",
            "Activity",
            "Event",
            "Event Date",
            "Members",
            "Registered At",
        ]
    )
    for team in TeamRegistration.objects.select_related(
        "leader", "activity", "activity__event"
    ).filter(activity__event=event):
        ws4.append(
            [
                team.team_name,
                team.leader.username,
                team.activity.name,
                team.activity.event.event,
                team.activity.event.date_of_event.isoformat()
                if team.activity.event.date_of_event
                else "",
                team.member_names,
                team.registered_at.isoformat(),
            ]
        )

    if ws4.max_row == 1:
        ws4.append(["No team registrations yet.", "", "", "", "", "", ""]) 

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"event_{event.id}_export.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=\"{filename}\""
    return response


@login_required
def export_role_reports_excel(request):
    """Export role-scoped Excel reports (participants, activities, analytics, teams, leaderboard)."""
    role = get_user_role(request.user)
    if role not in {"organizer", "coordinator"} and not request.user.is_superuser:
        return HttpResponse("You are not allowed to export reports.", status=403)

    # Events the user can see
    if request.user.is_superuser:
        events_qs = Event.objects.all()
    elif role == "organizer":
        events_qs = Event.objects.filter(user=request.user)
    else:
        event_ids = set(
            EventCoordinator.objects.filter(user=request.user).values_list("event_id", flat=True)
        )
        activity_event_ids = ActivityCoordinator.objects.filter(user=request.user).values_list(
            "activity__event_id", flat=True
        )
        event_ids.update(activity_event_ids)
        events_qs = Event.objects.filter(id__in=event_ids)

    accessible_event_ids = list(events_qs.values_list("id", flat=True))

    # Activities the user can see
    if request.user.is_superuser or role == "organizer":
        activities_qs = Activity.objects.filter(event_id__in=accessible_event_ids)
    else:
        activity_ids = set(
            ActivityCoordinator.objects.filter(user=request.user).values_list("activity_id", flat=True)
        )
        inherited_activity_ids = Activity.objects.filter(event_id__in=accessible_event_ids).values_list(
            "id", flat=True
        )
        activity_ids.update(inherited_activity_ids)
        activities_qs = Activity.objects.filter(id__in=activity_ids, event_id__in=accessible_event_ids)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0f172a")
    header_font = Font(bold=True, color="FFFFFF")

    # Participant List
    ws = wb.active
    ws.title = "Participant List"
    ws.append(["Name", "Email", "Phone", "Activity", "Status", "Team", "Attendance", "Score"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    attendance_map = {
        row.registration_id: row.status
        for row in AttendanceRecord.objects.filter(registration__activity__in=activities_qs)
    }
    score_map = {
        (row.activity_id, row.participant_id): row.score
        for row in LeaderboardEntry.objects.filter(activity__in=activities_qs)
    }
    team_map = {
        (row.activity_id, row.leader_id): row.team_name
        for row in TeamRegistration.objects.filter(activity__in=activities_qs)
    }
    for reg in ActivityRegistration.objects.select_related("participant", "activity", "activity__event").filter(
        activity__in=activities_qs
    ):
        full_name = reg.participant.get_full_name() or reg.participant.username
        team_name = team_map.get((reg.activity_id, reg.participant_id), "")
        ws.append(
            [
                full_name,
                reg.participant.email,
                getattr(reg.participant, "phone_number", "") or "",
                f"{reg.activity.name} ({reg.activity.event.event})",
                reg.status.title(),
                team_name or "-",
                attendance_map.get(reg.id, "Not recorded").title() if attendance_map.get(reg.id) else "Not recorded",
                score_map.get((reg.activity_id, reg.participant_id), 0),
            ]
        )

    # Activity Report
    ws_act = wb.create_sheet("Activity Report")
    ws_act.append(["Activity Name", "Total Registered", "Approved", "Rejected", "Attended"])
    for cell in ws_act[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for activity in activities_qs.select_related("event"):
        regs = ActivityRegistration.objects.filter(activity=activity)
        total = regs.count()
        approved = regs.filter(status=ActivityRegistration.STATUS_APPROVED).count()
        rejected = regs.filter(status=ActivityRegistration.STATUS_REJECTED).count()
        attended = AttendanceRecord.objects.filter(registration__activity=activity, status=AttendanceRecord.STATUS_PRESENT).count()
        ws_act.append([activity.name, total, approved, rejected, attended])

    # Event Analytics
    ws_evt = wb.create_sheet("Event Analytics")
    ws_evt.append(["Event Name", "Total Activities", "Total Registrations", "Total Attendance"])
    for cell in ws_evt[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    event_reg_counts = {
        row["event_id"]: row["count"]
        for row in EventRegistration.objects.filter(event_id__in=accessible_event_ids)
        .values("event_id")
        .annotate(count=models.Count("id"))
    }
    attendance_counts = {
        row["registration__activity__event_id"]: row["count"]
        for row in AttendanceRecord.objects.filter(registration__activity__event_id__in=accessible_event_ids)
        .values("registration__activity__event_id")
        .annotate(count=models.Count("id"))
    }
    for event in events_qs:
        ws_evt.append(
            [
                event.event,
                event.activities.count(),
                event_reg_counts.get(event.id, 0),
                attendance_counts.get(event.id, 0),
            ]
        )

    # Team Report
    ws_team = wb.create_sheet("Team Report")
    ws_team.append(["Team Name", "Leader", "Members", "Activity", "Status"])
    for cell in ws_team[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for team in TeamRegistration.objects.select_related("activity", "leader", "activity__event").filter(
        activity__in=activities_qs
    ):
        ws_team.append(
            [
                team.team_name,
                team.leader.get_full_name() or team.leader.username,
                team.member_names or "-",
                f"{team.activity.name} ({team.activity.event.event})",
                "Registered",
            ]
        )

    # Leaderboard
    ws_lb = wb.create_sheet("Leaderboard")
    ws_lb.append(["Rank", "Name", "Score", "Activity"])
    for cell in ws_lb[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    rows = []
    for entry in LeaderboardEntry.objects.select_related("participant", "activity", "activity__event").filter(
        activity__in=activities_qs
    ):
        rows.append(
            (entry.score, entry.participant.get_full_name() or entry.participant.username, entry.activity)
        )
    rows.sort(key=lambda r: r[0], reverse=True)
    for idx, row in enumerate(rows, start=1):
        score, name, activity = row
        ws_lb.append([idx, name, score, f"{activity.name} ({activity.event.event})"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = timezone.now().strftime("%Y%m%d%H%M")
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="ems_reports_{timestamp}.xlsx"'
    return response


@login_required
def manage_activities(request):
    is_head = _is_head_coordinator(request.user)
    is_event_coord = _is_event_coordinator(request.user)
    is_activity_coord = _is_activity_coordinator(request.user)
    can_manage = (
        _is_organizer(request.user)
        or request.user.is_staff
        or request.user.is_superuser
        or is_head
        or is_event_coord
    )
    selected_event_id = request.GET.get("event_id") or request.POST.get("event_id")
    can_view_event_names = (
        request.user.is_superuser
        or request.user.is_staff
        or _is_organizer(request.user)
        or is_head
        or is_event_coord
    )

    # Provide all events for dropdown
    if request.user.is_superuser or request.user.is_staff:
        events = Event.objects.all().order_by("-id")
    elif _is_organizer(request.user):
        events = Event.objects.filter(user=request.user).order_by("-id")
    else:
        events = Event.objects.none()

    # Provide all activities for the selected event
    activities = []
    if selected_event_id:
        activities = Activity.objects.filter(event_id=selected_event_id).order_by("start_time")

    # Users available to assign as coordinators
    users = get_user_model().objects.filter(is_active=True).order_by("username")
    # Collect all form fields
    form_data = {
        "event_id": request.POST.get("event_id", selected_event_id or ""),
        "activity_name": request.POST.get("activity_name", ""),
        "description": request.POST.get("description", ""),
        "rules": request.POST.get("rules", ""),
        "eligibility": request.POST.get("eligibility", ""),
        "date": request.POST.get("date", ""),
        "start_time": request.POST.get("start_time", ""),
        "end_time": request.POST.get("end_time", ""),
        "prize": request.POST.get("prize", ""),
        "max_participants": request.POST.get("max_participants", ""),
        "is_team_event": request.POST.get("is_team_event", None),
        "team_size": request.POST.get("team_size", ""),
        # Comma/newline-separated coordinator emails entered by the organizer
        "coordinator_emails": request.POST.get("coordinator_emails", ""),
        "registration_fee": request.POST.get("registration_fee", ""),
        "registration_form_fields": request.POST.get("registration_form_fields", ""),
        # Handle file upload
        "registration_form": request.FILES.get("registration_form") if request.method == "POST" else None,
    }

    if request.method == "POST":
        # Debug output: print all POST and FILES data to console/log (and coordinators list)
        print("[DEBUG] POST data:", dict(request.POST))
        print("[DEBUG] FILES data:", dict(request.FILES))
        print("[DEBUG] coordinator_emails:", request.POST.get("coordinator_emails"))
        # Optionally, you can log this to a file or use logging module

        if not can_manage:
            messages.error(request, "Only organizers can create activities.")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })

        # Validate required fields
        # Check all required fields, including date and file if needed
        # Coordinator emails (comma/newline-separated) are required
        required_fields = ["event_id", "activity_name", "registration_fee", "start_time", "end_time", "date", "coordinator_emails"]
        missing = [f for f in required_fields if not form_data.get(f)]
        if missing:
            messages.error(request, f"Missing required fields: {', '.join(missing)}. Please fill all required fields.")
            print(f"[DEBUG] Missing fields: {missing}")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })

        # Validate times
        import datetime
        time_format = "%H:%M"
        try:
            start_time = datetime.datetime.strptime(form_data["start_time"], time_format).time()
        except ValueError:
            messages.error(request, "Start time must be in HH:MM format (e.g., 14:30).")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })
        try:
            end_time = datetime.datetime.strptime(form_data["end_time"], time_format).time()
        except ValueError:
            messages.error(request, "End time must be in HH:MM format (e.g., 16:00).")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })

        # Validate team event
        is_team_event = form_data["is_team_event"] == "on"
        team_size_value = form_data.get("team_size")
        team_size = None
        if is_team_event:
            if not team_size_value:
                messages.error(request, "Team size is required for team events.")
                return render(request, "events/manage_activities.html", {
                    "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                    "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                    "users": users,
                })
            try:
                team_size = int(team_size_value)
            except ValueError:
                messages.error(request, "Team size must be a number.")
                return render(request, "events/manage_activities.html", {
                    "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                    "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                    "users": users,
                })
            if team_size < 2:
                messages.error(request, "Team size must be at least 2.")
                return render(request, "events/manage_activities.html", {
                    "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                    "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                    "users": users,
                })

        # Parse coordinator emails (split by comma, semicolon, or newline)
        raw_emails = (form_data.get("coordinator_emails") or "")
        potential = [e.strip() for e in re.split(r"[\n,;]+", raw_emails) if e.strip()]
        if not potential:
            messages.error(request, "Please provide at least one coordinator email.")
            print("[DEBUG] No coordinator emails provided")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })
        # Validate each email
        invalid = []
        valid_emails = []
        for em in potential:
            try:
                validate_email(em)
                valid_emails.append(em.lower())
            except ValidationError:
                invalid.append(em)
        if invalid:
            messages.error(request, f"Invalid coordinator email(s): {', '.join(invalid)}")
            print(f"[DEBUG] Invalid emails: {invalid}")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })

        event = get_object_or_404(Event, id=form_data["event_id"])
        if not _can_manage_event_activities(request.user, event):
            messages.error(request, "You do not have permission to add activities for this event.")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })

        activity, created = Activity.objects.get_or_create(
            event=event,
            name=form_data["activity_name"],
            defaults={
                "registration_fee": form_data["registration_fee"],
                "start_time": start_time,
                "end_time": end_time,
                "is_team_event": is_team_event,
                "team_size": team_size,
                "description": form_data["description"],
                "rules": form_data["rules"],
                "eligibility": form_data["eligibility"],
                "prize": form_data["prize"],
                "date": form_data["date"] if form_data["date"] else None,
                "max_participants": int(form_data["max_participants"]) if form_data["max_participants"] else None,
            },
        )
        if not created:
            messages.info(request, "This activity already exists for the event.")
            return render(request, "events/manage_activities.html", {
                "form_data": form_data, "can_manage": can_manage, "selected_event_id": selected_event_id,
                "can_view_event_names": can_view_event_names, "events": events, "activities": activities,
                "users": users,
            })

        # Create invites for each valid email
        for email in valid_emails:
            invite, created_invite = EventCoordinatorInvite.objects.get_or_create(
                event=event,
                email=email,
                status=EventCoordinatorInvite.STATUS_PENDING,
                defaults={"invited_by": request.user},
            )
            if not created_invite:
                messages.info(request, "Coordinator already invited for this event.")
            else:
                _send_event_invite_email(request, invite)

        messages.success(request, "Activity created successfully.")
        return redirect("events:manage_activities")

    # For GET or after error, show form with preserved data and all context
    return render(request, "events/manage_activities.html", {
        "form_data": form_data,
        "can_manage": can_manage,
        "selected_event_id": selected_event_id,
        "can_view_event_names": can_view_event_names,
        "events": events,
        "activities": activities,
        "users": users,
    })


@login_required
def event_dashboard(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Use centralized permission check as requested
    can_manage = _can_manage_event_activities(request.user, event)
    can_view_participants = _can_view_event_participants(request.user, event)
    
    if not (can_manage or can_view_participants):
        messages.error(request, "You do not have permission to access this event dashboard.")
        return HttpResponse("Forbidden: You do not have permission to access this dashboard.", status=403)
    can_invite_event_coordinator = (
        request.user.is_superuser
        or request.user.is_staff
        or _is_head_coordinator(request.user)
        or (_is_organizer(request.user) and event.user_id == request.user.id)
    )

    if request.method == "POST":
        action = request.POST.get("action") or "add_activity"

        if action == "invite_event_coordinator":
            if not can_invite_event_coordinator:
                messages.error(request, "You do not have permission to invite event coordinators.")
                return redirect("events:event_dashboard", event_id=event_id)

            coordinator_email = (request.POST.get("event_coordinator_email") or "").strip().lower()
            if not coordinator_email:
                messages.error(request, "Coordinator email is required.")
                return redirect("events:event_dashboard", event_id=event_id)

            try:
                validate_email(coordinator_email)
            except ValidationError:
                messages.error(request, "Please enter a valid coordinator email.")
                return redirect("events:event_dashboard", event_id=event_id)

            invite, created_invite = EventCoordinatorInvite.objects.get_or_create(
                event=event,
                email=coordinator_email,
                status=EventCoordinatorInvite.STATUS_PENDING,
                defaults={"invited_by": request.user},
            )
            if not created_invite:
                messages.info(request, "Coordinator already invited for this event.")
            else:
                _send_event_invite_email(request, invite)

            messages.success(request, "Event coordinator invite sent.")
            return redirect("events:event_dashboard", event_id=event_id)

        if not can_manage:
            messages.error(request, "Only authorized coordinators can add activities.")
            return redirect("events:event_dashboard", event_id=event_id)

        activity_name = (request.POST.get("activity_name") or "").strip()
        registration_fee = (request.POST.get("registration_fee") or "").strip()
        start_time = (request.POST.get("start_time") or "").strip()
        end_time = (request.POST.get("end_time") or "").strip()
        coordinator_email = (request.POST.get("coordinator_email") or "").strip().lower()

        if not activity_name or not registration_fee or not start_time or not end_time or not coordinator_email:
            messages.error(request, "Activity name, time, fee, and coordinator email are required.")
            return redirect("events:event_dashboard", event_id=event_id)

        is_team_event = request.POST.get("is_team_event") == "on"
        team_size_value = request.POST.get("team_size")
        team_size = None
        if is_team_event:
            if not team_size_value:
                messages.error(request, "Team size is required for team events.")
                return redirect("events:event_dashboard", event_id=event_id)
            try:
                team_size = int(team_size_value)
            except ValueError:
                messages.error(request, "Team size must be a number.")
                return redirect("events:event_dashboard", event_id=event_id)
            if team_size < 2:
                messages.error(request, "Team size must be at least 2.")
                return redirect("events:event_dashboard", event_id=event_id)

        try:
            validate_email(coordinator_email)
        except ValidationError:
            messages.error(request, "Please enter a valid coordinator email.")
            return redirect("events:event_dashboard", event_id=event_id)

        activity, created = Activity.objects.get_or_create(
            event=event,
            name=activity_name,
            defaults={
                "registration_fee": registration_fee,
                "start_time": start_time,
                "end_time": end_time,
                "is_team_event": is_team_event,
                "team_size": team_size,
            },
        )
        if not created:
            messages.info(request, "This activity already exists for the event.")
            return redirect("events:event_dashboard", event_id=event_id)

        invite, created_invite = EventCoordinatorInvite.objects.get_or_create(
            event=event,
            email=coordinator_email,
            status=EventCoordinatorInvite.STATUS_PENDING,
            defaults={"invited_by": request.user},
        )
        if not created_invite:
            messages.info(request, "Coordinator already invited for this event.")
        else:
            _send_event_invite_email(request, invite)

        messages.success(request, "Activity created successfully.")
        return redirect("events:event_dashboard", event_id=event_id)

    activities = Activity.objects.select_related("event").filter(event=event).order_by("-id")
    activity_coordinators = ActivityCoordinator.objects.select_related("user", "activity")
    activity_coordinator_map = {}
    for row in activity_coordinators:
        activity_coordinator_map.setdefault(row.activity_id, []).append(row.user)

    registration_counts = {
        row["activity_id"]: row["count"]
        for row in ActivityRegistration.objects.filter(activity__event=event)
        .values("activity_id")
        .annotate(count=models.Count("id"))
    }

    activity_rows = []
    for activity in activities:
        coordinators = activity_coordinator_map.get(activity.id, [])
        coordinator_names = ", ".join(user.username for user in coordinators) or "-"
        activity_rows.append(
            {
                "activity": activity,
                "coordinator_names": coordinator_names,
                "registration_count": registration_counts.get(activity.id, 0),
            }
        )

    can_view_participants = _can_view_event_participants(request.user, event)
    if can_view_participants:
        event_registrations = EventRegistration.objects.select_related(
            "participant"
        ).filter(event=event).order_by("-registered_at")
        activity_registrations = ActivityRegistration.objects.select_related(
            "participant", "activity"
        ).filter(activity__event=event).order_by("-registered_at")
    else:
        event_registrations = EventRegistration.objects.none()
        activity_registrations = ActivityRegistration.objects.none()

    return render(
        request,
        "events/event_dashboard.html",
        {
            "event": event,
            "activities": activity_rows,
            "can_manage": can_manage,
            "can_invite_event_coordinator": can_invite_event_coordinator,
            "can_view_participants": can_view_participants,
            "event_registrations": event_registrations,
            "activity_registrations": activity_registrations,
            "event_coordinators": EventCoordinator.objects.select_related("user").filter(event=event),
        },
    )


@login_required
def delete_event(request, event_id):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    event = get_object_or_404(Event, id=event_id)

    if not _can_manage_event(request.user, event):
        messages.error(request, "You do not have permission to delete this event.")
        return redirect("events:manage_events")

    event.delete()
    messages.success(request, "Event deleted successfully.")
    return redirect("events:manage_events")


@login_required
def accept_activity_invite(request, token):
    invite = get_object_or_404(EventCoordinatorInvite, token=token)

    if invite.status == EventCoordinatorInvite.STATUS_ACCEPTED:
        messages.info(request, "This invite has already been accepted.")
        return redirect("events:dashboard")

    if not request.user.email:
        messages.error(request, "Your account must have an email to accept invites.")
        return redirect("events:dashboard")

    if request.user.email.lower() != invite.email.lower():
        messages.error(request, "This invite is for a different email address.")
        return redirect("events:dashboard")

    if not Profile.objects.filter(user=request.user, role="coordinator").exists():
        messages.error(request, "Only coordinator accounts can accept this invite.")
        return redirect("events:dashboard")

    ActivityCoordinator.objects.get_or_create(activity=invite.activity, user=request.user)
    invite.mark_accepted()
    messages.success(request, "You are now a coordinator for this activity.")
    return redirect("events:dashboard")


@login_required
def accept_event_invite(request, token):
    invite = get_object_or_404(EventCoordinatorInvite, token=token)

    if invite.status == EventCoordinatorInvite.STATUS_ACCEPTED:
        messages.info(request, "This invite has already been accepted.")
        return redirect("events:dashboard")

    if not request.user.email:
        messages.error(request, "Your account must have an email to accept invites.")
        return redirect("events:dashboard")

    if request.user.email.lower() != invite.email.lower():
        messages.error(request, "This invite is for a different email address.")
        return redirect("events:dashboard")

    if not Profile.objects.filter(user=request.user, role="coordinator").exists():
        messages.error(request, "Only coordinator accounts can accept this invite.")
        return redirect("events:dashboard")

    EventCoordinator.objects.get_or_create(event=invite.event, user=request.user)
    invite.mark_accepted()
    messages.success(request, "You are now a coordinator for this event.")
    return redirect("events:dashboard")


@login_required
def scan_attendance(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    qr_token = (request.POST.get("qr_token") or "").strip()
    if not qr_token:
        messages.error(request, "QR token is required.")
        return redirect("events:dashboard")

    registration = get_object_or_404(ActivityRegistration, qr_token=qr_token)
    if not _can_manage_activity(request.user, registration.activity):
        messages.error(request, "You do not have permission to mark attendance.")
        return redirect("events:dashboard")

    attendance, _ = AttendanceRecord.objects.get_or_create(registration=registration)
    attendance.status = AttendanceRecord.STATUS_PRESENT
    attendance.scanned_by = request.user
    attendance.save(update_fields=["status", "scanned_by", "scanned_at"])

    messages.success(request, "Attendance recorded.")
    return redirect("events:dashboard")


@login_required
def update_leaderboard(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    activity_id = request.POST.get("activity_id")
    participant_id = request.POST.get("participant_id")
    score_value = request.POST.get("score")

    if not activity_id or not participant_id or score_value is None:
        messages.error(request, "Activity, participant, and score are required.")
        return redirect("events:dashboard")

    activity = get_object_or_404(Activity, id=activity_id)
    if not _can_manage_activity(request.user, activity):
        messages.error(request, "You do not have permission to update the leaderboard.")
        return redirect("events:dashboard")

    try:
        score_int = int(score_value)
    except ValueError:
        messages.error(request, "Score must be a number.")
        return redirect("events:dashboard")

    entry, _ = LeaderboardEntry.objects.get_or_create(
        activity=activity,
        participant_id=participant_id,
        defaults={"score": score_int, "updated_by": request.user},
    )
    if entry.score != score_int or entry.updated_by_id != request.user.id:
        entry.score = score_int
        entry.updated_by = request.user
        entry.save(update_fields=["score", "updated_by", "updated_at"])

    messages.success(request, "Leaderboard updated.")
    return redirect("events:dashboard")


@login_required
def issue_certificate(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    activity_id = request.POST.get("activity_id")
    participant_id = request.POST.get("participant_id")

    if not activity_id or not participant_id:
        messages.error(request, "Activity and participant are required.")
        return redirect("events:dashboard")

    activity = get_object_or_404(Activity, id=activity_id)
    if not _can_manage_activity(request.user, activity):
        messages.error(request, "You do not have permission to issue certificates.")
        return redirect("events:dashboard")

    certificate, created = Certificate.objects.get_or_create(
        activity=activity,
        participant_id=participant_id,
        defaults={"issued_by": request.user},
    )
    if not created and certificate.issued_by_id != request.user.id:
        certificate.issued_by = request.user
        certificate.save(update_fields=["issued_by"])

    messages.success(request, "Certificate issued.")
    return redirect("events:dashboard")


# ============================================================================
# COORDINATOR: Approve / Reject Registration
# ============================================================================

@login_required
def approve_registration(request, registration_id):
    """Coordinator approves an activity registration."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    registration = get_object_or_404(ActivityRegistration, id=registration_id)
    if not _can_manage_activity(request.user, registration.activity):
        return HttpResponse("Forbidden", status=403)
    registration.status = ActivityRegistration.STATUS_APPROVED
    registration.save(update_fields=["status"])
    messages.success(request, f"Registration for {registration.participant.username} approved.")
    return redirect(request.POST.get("next") or "events:coordinator_dashboard")


@login_required
def reject_registration(request, registration_id):
    """Coordinator rejects an activity registration."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    registration = get_object_or_404(ActivityRegistration, id=registration_id)
    if not _can_manage_activity(request.user, registration.activity):
        return HttpResponse("Forbidden", status=403)
    registration.status = ActivityRegistration.STATUS_REJECTED
    registration.save(update_fields=["status"])
    messages.warning(request, f"Registration for {registration.participant.username} rejected.")
    return redirect(request.POST.get("next") or "events:coordinator_dashboard")


# ============================================================================
# COORDINATOR: QR Scanner Page (GET → renders camera scanner UI)
# ============================================================================

@login_required
def qr_scanner_page(request):
    """Renders the camera-based QR scanner UI for coordinators."""
    if not _is_coordinator(request.user) and not request.user.is_superuser and not request.user.is_staff:
        return HttpResponse("Forbidden", status=403)
    return render(request, "coordinator/qr_scanner.html")


# ============================================================================
# COORDINATOR: Participant List for an Activity
# ============================================================================

@login_required
def activity_participants(request, activity_id):
    """Coordinator views all participants registered for their activity."""
    activity = get_object_or_404(Activity, id=activity_id)
    if not _can_manage_activity(request.user, activity):
        return HttpResponse("Forbidden", status=403)

    registrations = (
        ActivityRegistration.objects
        .filter(activity=activity)
        .select_related("participant")
        .prefetch_related("attendance")
        .order_by("-registered_at")
    )
    attendance_map = {r.registration_id: r for r in AttendanceRecord.objects.filter(registration__activity=activity)}
    certificate_map = {c.participant_id: c for c in Certificate.objects.filter(activity=activity)}
    leaderboard_map = {e.participant_id: e for e in LeaderboardEntry.objects.filter(activity=activity)}

    rows = []
    for reg in registrations:
        rows.append({
            "registration": reg,
            "attendance": attendance_map.get(reg.id),
            "certificate": certificate_map.get(reg.participant_id),
            "leaderboard": leaderboard_map.get(reg.participant_id),
        })

    return render(request, "coordinator/activity_participants.html", {
        "activity": activity,
        "rows": rows,
    })


# ============================================================================
# PARTICIPANT: Download Certificate (PDF)
# ============================================================================

@login_required
@require_GET
def download_certificate(request, certificate_id):
    """Participant downloads their certificate as PDF."""
    from participant.models import Certificate as CertModel
    cert = get_object_or_404(CertModel, certificate_id=certificate_id, participant=request.user)

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4

    # Background
    c.setFillColorRGB(0.05, 0.09, 0.18)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    # Gold border
    c.setStrokeColorRGB(0.96, 0.77, 0.16)
    c.setLineWidth(4)
    c.rect(20, 20, width - 40, height - 40, fill=0, stroke=1)

    # Title
    c.setFillColorRGB(0.96, 0.77, 0.16)
    c.setFont("Helvetica-Bold", 32)
    c.drawCentredString(width / 2, height - 100, "Certificate of Participation")

    # Subtitle
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica", 16)
    c.drawCentredString(width / 2, height - 145, "This is to certify that")

    # Name
    c.setFillColorRGB(0.96, 0.77, 0.16)
    c.setFont("Helvetica-Bold", 26)
    full_name = cert.participant.get_full_name() or cert.participant.username
    c.drawCentredString(width / 2, height - 195, full_name)

    # Body
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica", 14)
    c.drawCentredString(width / 2, height - 235, f"successfully participated in")
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, height - 265, cert.activity.name)
    c.setFont("Helvetica", 12)
    c.drawCentredString(width / 2, height - 295, f"— {cert.activity.event.event} —")

    # Date
    c.setFillColorRGB(0.6, 0.6, 0.6)
    c.setFont("Helvetica", 11)
    c.drawCentredString(width / 2, height - 340, f"Issued on: {cert.issued_at.strftime('%B %d, %Y')}")

    # Certificate ID
    c.setFont("Helvetica", 9)
    c.drawCentredString(width / 2, 50, f"Certificate ID: {cert.certificate_id}")

    c.showPage()
    c.save()
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="certificate_{cert.certificate_id}.pdf"'
    return response


@login_required
@require_GET
def registration_summary_pdf(request, registration_id):
    """PDF of submitted registration data; participants see their own, organizers/coordinators see their event data."""
    registration = get_object_or_404(EventRegistration, id=registration_id)
    event = registration.event

    can_view = (
        registration.participant_id == request.user.id
        or request.user.is_superuser
        or _can_manage_event(request.user, event)
        or _is_event_coordinator_for_event(request.user, event)
    )
    if not can_view:
        return HttpResponse("You do not have permission to view this registration.", status=403)

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 18)
    c.drawString(30 * mm, height - 35 * mm, "Registration Summary")
    c.setFont("Helvetica", 12)
    c.drawString(30 * mm, height - 50 * mm, f"Event: {event.event}")
    c.drawString(30 * mm, height - 60 * mm, f"Participant: {registration.participant.get_full_name() or registration.participant.username}")
    c.drawString(30 * mm, height - 70 * mm, f"Registered at: {registration.registered_at.strftime('%Y-%m-%d %H:%M')}")

    y = height - 90 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30 * mm, y, "Submitted Data")
    y -= 10 * mm
    c.setFont("Helvetica", 11)
    data = registration.form_data or {}
    if not data:
        c.drawString(30 * mm, y, "No additional data submitted.")
    else:
        for key, value in data.items():
            c.drawString(30 * mm, y, f"{key}: {value}")
            y -= 8 * mm
            if y < 30 * mm:
                c.showPage()
                y = height - 30 * mm

    c.showPage()
    c.save()
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=\"registration_{registration.id}.pdf\"'
    return response


@login_required
@require_GET
def event_report_pdf(request, event_id):
    """Compact event report for organizers/coordinators with stats and winners."""
    event = get_object_or_404(Event, id=event_id)
    can_view = (
        request.user.is_superuser
        or _can_manage_event(request.user, event)
        or _is_event_coordinator_for_event(request.user, event)
    )
    if not can_view:
        return HttpResponse("You do not have permission to view this event.", status=403)

    activities = Activity.objects.filter(event=event)
    reg_count = EventRegistration.objects.filter(event=event).count()
    act_reg_count = ActivityRegistration.objects.filter(activity__event=event).count()
    attendance_count = AttendanceRecord.objects.filter(registration__activity__event=event, status=AttendanceRecord.STATUS_PRESENT).count()

    # winners: highest score per activity
    winners = []
    for activity in activities:
        top = (
            LeaderboardEntry.objects.filter(activity=activity)
            .select_related("participant")
            .order_by("-score", "participant__username")
            .first()
        )
        if top:
            winners.append((activity.name, top.participant.get_full_name() or top.participant.username, top.score))

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 20)
    c.drawString(30 * mm, height - 35 * mm, "Event Report")
    c.setFont("Helvetica", 12)
    c.drawString(30 * mm, height - 48 * mm, f"Event: {event.event}")
    c.drawString(30 * mm, height - 58 * mm, f"Date: {event.date_of_event or 'TBA'}")
    c.drawString(30 * mm, height - 68 * mm, f"Venue: {event.venue or 'TBA'}")

    c.setFont("Helvetica-Bold", 12)
    c.drawString(30 * mm, height - 88 * mm, "Analytics")
    c.setFont("Helvetica", 11)
    c.drawString(30 * mm, height - 98 * mm, f"Activities: {activities.count()}")
    c.drawString(30 * mm, height - 108 * mm, f"Event Registrations: {reg_count}")
    c.drawString(30 * mm, height - 118 * mm, f"Activity Registrations: {act_reg_count}")
    c.drawString(30 * mm, height - 128 * mm, f"Attendance Marked: {attendance_count}")

    y = height - 150 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30 * mm, y, "Winners / Leaderboard Tops")
    y -= 10 * mm
    c.setFont("Helvetica", 11)
    if winners:
        for activity_name, participant_name, score in winners:
            c.drawString(30 * mm, y, f"{activity_name}: {participant_name} (Score: {score})")
            y -= 8 * mm
            if y < 30 * mm:
                c.showPage()
                y = height - 30 * mm
    else:
        c.drawString(30 * mm, y, "No leaderboard data yet.")

    c.showPage()
    c.save()
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=\"event_report_{event.id}.pdf\"'
    return response


# ============================================================================
# PARTICIPANT: Activity Gate Pass (QR) for Activity Registration
# ============================================================================

@login_required
@require_GET
def generate_activity_gate_pass(request, registration_id):
    """Participant generates gate pass PDF for an approved activity registration."""
    registration = get_object_or_404(ActivityRegistration, id=registration_id, participant=request.user)
    if registration.status != ActivityRegistration.STATUS_APPROVED:
        messages.error(request, "Your registration must be approved before you can download a gate pass.")
        return redirect("events:participant_dashboard")

    activity = registration.activity
    event = activity.event

    qr_data = f"ACT-REG-{registration.id}|TOKEN-{registration.qr_token}|ACT-{activity.id}|EVENT-{event.id}"
    qr_img = qrcode.make(qr_data)
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 20)
    c.drawString(30 * mm, height - 40 * mm, f"Activity Gate Pass")
    c.setFont("Helvetica-Bold", 14)
    c.drawString(30 * mm, height - 55 * mm, f"Event: {event.event}")
    c.setFont("Helvetica", 12)
    c.drawString(30 * mm, height - 68 * mm, f"Activity: {activity.name}")
    c.drawString(30 * mm, height - 80 * mm, f"Participant: {request.user.get_full_name() or request.user.username}")
    c.drawString(30 * mm, height - 92 * mm, f"Date: {event.date_of_event}")
    c.drawString(30 * mm, height - 104 * mm, f"Venue: {event.venue}")
    c.drawInlineImage(qr_buffer, 30 * mm, height - 165 * mm, 55 * mm, 55 * mm)
    c.setFont("Helvetica", 9)
    c.drawString(30 * mm, height - 175 * mm, f"Token: {registration.qr_token}")
    c.showPage()
    c.save()
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="gate_pass_activity_{registration.id}.pdf"'
    return response



# Organizer Dashboard (superuser only)
@login_required
def organizer_dashboard(request):
    """Organizer dashboard - shows their created events."""
    user_role = get_user_role(request.user)
    if user_role != 'organizer' and not request.user.is_superuser:
        messages.error(request, "Only organizers can access this dashboard.")
        return redirect('events:user')
    
    is_head = _is_head_coordinator(request.user)
    is_event_coord = _is_event_coordinator(request.user)
    is_activity_coord = _is_activity_coordinator(request.user)
    can_view_participant_map = _is_organizer(request.user) or is_head or is_event_coord
    can_register = _is_participant(request.user)
    can_view_event_dashboard = request.user.is_superuser or _is_organizer(request.user)
    can_check_payments = (
        _is_coordinator(request.user)
        or can_view_participant_map
        or request.user.is_staff
        or request.user.is_superuser
    )

    # Show only events created by this organizer
    if request.user.is_superuser:
        events = Event.objects.select_related("user").all().order_by("-id")
    else:
        events = Event.objects.filter(user=request.user).select_related("user").order_by("-id")
    
    event_activity_rows = []
    if can_view_event_dashboard:
        # Optimization: use prefetch_related for activities and count registrations
        if request.user.is_superuser:
            qs = Event.objects.prefetch_related("activities", "event_registrations").order_by("-id")
        else:
            qs = Event.objects.filter(user=request.user).prefetch_related("activities", "event_registrations").order_by("-id")
        
        for event in qs:
            activity_names = ", ".join(activity.name for activity in event.activities.all())
            participant_count = event.event_registrations.count()
            event_activity_rows.append({
                "event": event,
                "activity_names": activity_names or "-",
                "participant_count": participant_count,
            })

    # Organizer analytics
    total_events = events.count() if hasattr(events, 'count') else len(events)
    total_registrations = EventRegistration.objects.filter(event__user=request.user).count()
    # Registrations per activity
    regs_per_activity = {
        r['activity_id']: r['count']
        for r in ActivityRegistration.objects.filter(activity__event__user=request.user).values('activity_id').annotate(count=models.Count('id'))
    }
    # Visitor count: active anonymous sessions (no _auth_user_id in session data)
    try:
        active_sessions = Session.objects.filter(expire_date__gt=timezone.now())
        visitor_count = sum(1 for s in active_sessions if not s.get_decoded().get('_auth_user_id'))
    except Exception:
        visitor_count = 0

    return render(
        request,
        "organizer_dashboard.html",
        {
            "event_activity_rows": event_activity_rows,
            "events": events,
            "can_register": can_register,
            "can_view_participant_map": can_view_participant_map,
            "can_check_payments": can_check_payments,
            "analytics": {
                "total_events": total_events,
                "total_registrations": total_registrations,
                    "regs_per_activity": regs_per_activity,
                    "visitor_count": visitor_count,
            },
        },
    )


# Coordinator Dashboard (organizer or coordinator)
@login_required
def coordinator_dashboard(request):
    """Coordinator dashboard - shows assigned events and activities."""
    user_role = get_user_role(request.user)
    if user_role != 'coordinator' and not request.user.is_superuser:
        messages.error(request, "Only coordinators can access this dashboard.")
        return redirect('events:user')
    
    is_head = _is_head_coordinator(request.user)
    is_event_coord = _is_event_coordinator(request.user)
    is_activity_coord = _is_activity_coordinator(request.user)
    can_view_participant_map = _is_organizer(request.user) or is_head or is_event_coord
    can_register = _is_participant(request.user)
    can_view_event_dashboard = request.user.is_superuser or _is_coordinator(request.user)
    can_check_payments = (
        _is_coordinator(request.user)
        or can_view_participant_map
        or request.user.is_staff
        or request.user.is_superuser
    )

    # Coordinator logic: only assigned events and activities
    assigned_events = EventCoordinator.objects.filter(user=request.user).select_related("event", "event__user")
    assigned_activities = ActivityCoordinator.objects.filter(user=request.user).select_related("activity", "activity__event")

    event_activity_rows = []
    # Map events
    for ec in assigned_events:
        event = ec.event
        activity_names = ", ".join(act.name for act in event.activities.all())
        event_activity_rows.append({
            "event": event,
            "activity_names": activity_names or "-",
            "assignment_type": "Event",
            "participant_count": event.event_registrations.count(),
            "activities": list(event.activities.all()),
        })
    
    # Map specifically assigned activities if not already part of an assigned event
    assigned_event_ids = [ec.event_id for ec in assigned_events]
    for ac in assigned_activities:
        if ac.activity.event_id not in assigned_event_ids:
            event = ac.activity.event
            event_activity_rows.append({
                "event": event,
                "activity_names": ac.activity.name,
                "assignment_type": "Activity",
                "participant_count": ac.activity.activity_registrations.count(),
                "activities": [ac.activity],
            })

    return render(
        request,
        "coordinator_dashboard.html",
        {
            "event_activity_rows": event_activity_rows,
            "assigned_events": assigned_events,
            "can_register": can_register,
            "can_view_participant_map": can_view_participant_map,
            "can_check_payments": can_check_payments,
        },
    )


# Shared Dashboard (participants, organizers, and coordinators)
@login_required
def participant_dashboard(request):
    """Shared dashboard - shows the role-appropriate dashboard experience."""
    user_role = get_user_role(request.user)
    if (
        user_role not in {"participant", "organizer", "coordinator"}
        and not request.user.is_superuser
    ):
        messages.error(request, "Only participants, organizers, and coordinators can access this dashboard.")
        return redirect('accounts:role_selection')
    is_head = _is_head_coordinator(request.user)
    is_event_coord = _is_event_coordinator(request.user)
    can_view_participant_map = _is_organizer(request.user) or is_head or is_event_coord
    can_register = _is_participant(request.user)
    can_check_payments = (
        _is_coordinator(request.user)
        or can_view_participant_map
        or request.user.is_staff
        or request.user.is_superuser
    )

    # Participant logic: Joined vs Recommended
    joined_registrations = EventRegistration.objects.filter(participant=request.user).select_related("event", "event__user")
    joined_event_ids = joined_registrations.values_list("event_id", flat=True)
    
    joined_events = [reg.event for reg in joined_registrations]
    recommended_events = Event.objects.exclude(id__in=joined_event_ids).select_related("user").order_by("-id")[:6]
    featured_event = recommended_events[0] if recommended_events else (joined_events[0] if joined_events else None)

    # Activity Registrations Overhaul
    activity_registrations = ActivityRegistration.objects.filter(
        participant=request.user
    ).select_related("activity", "activity__event").order_by("-registered_at")

    available_activities = Activity.objects.select_related("event", "event__user").order_by("-created_at")[:8]
    upcoming_activities = Activity.objects.select_related("event", "event__user").filter(
        event__date_of_event__gte=timezone.now().date()
    ).order_by("event__date_of_event", "start_time")[:4]

    leaderboard_preview = (
        LeaderboardEntry.objects.select_related("activity", "participant", "activity__event")
        .order_by("-score", "-updated_at")[:5]
    )

    # We need attendance and certificate for these activities
    activity_ids = [reg.activity_id for reg in activity_registrations]
    attendance_map = {
        row.registration_id: row for row in AttendanceRecord.objects.filter(registration__in=activity_registrations)
    }
    certificate_map = {
        row.activity_id: row for row in Certificate.objects.filter(participant=request.user, activity_id__in=activity_ids)
    }

    activity_rows = []
    for reg in activity_registrations:
        activity_rows.append({
            "registration": reg,
            "attendance": attendance_map.get(reg.id),
            "certificate": certificate_map.get(reg.activity_id)
        })

    return render(
        request,
        "participant_dashboard.html",
        {
            "joined_events": joined_events,
            "recommended_events": recommended_events,
            "featured_event": featured_event,
            "available_activities": available_activities,
            "upcoming_activities": upcoming_activities,
            "leaderboard_preview": leaderboard_preview,
            "activity_rows": activity_rows,
            "can_register": can_register,
            "can_view_participant_map": can_view_participant_map,
            "can_check_payments": can_check_payments,
        },
    )

# ============================================================================
# NEW VIEWS FOR EVENT MANAGEMENT SYSTEM
# ============================================================================

# Note: Redundant create_event and create_activity views removed.
# Use manage_events and manage_activities Views instead.


def event_website_home(request, event_id):
    """Public-facing event website homepage"""
    event = get_object_or_404(Event, id=event_id)
    
    if event.date_of_event:
        today = timezone.now().date()
        days_left = (event.date_of_event - today).days
    else:
        days_left = None
    
    context = {
        'event': event,
        'days_left': days_left,
        'activities_count': event.activities.count(),
    }
    return render(request, 'website/event_home.html', context)


def event_website_activities(request, event_id):
    """Public activities page"""
    event = get_object_or_404(Event, id=event_id)
    # annotate each activity with registration counts and remaining slots
    activities_qs = event.activities.all().order_by('start_time')
    # get counts in a single query
    from participant.models import ActivityRegistration
    reg_counts = {
        r['activity_id']: r['count']
        for r in ActivityRegistration.objects.filter(activity__event=event).values('activity_id').annotate(count=models.Count('id'))
    }

    activities = []
    user_registered_activity_ids = set()
    if request.user.is_authenticated:
        user_registered_activity_ids = set(
            ActivityRegistration.objects.filter(participant=request.user, activity__event=event).values_list('activity_id', flat=True)
        )

    for act in activities_qs:
        current = reg_counts.get(act.id, 0)
        maxp = act.max_participants or None
        remaining = None if maxp is None else max(0, maxp - current)
        percent = None
        if maxp:
            percent = int((current / maxp) * 100) if maxp > 0 else 100
        activities.append({
            'obj': act,
            'current_registrations': current,
            'max_participants': maxp,
            'remaining_slots': remaining,
            'fill_percent': percent,
            'user_registered': act.id in user_registered_activity_ids,
        })

    context = {
        'event': event,
        'activities': activities,
    }
    return render(request, 'website/event_activities.html', context)


def event_website_schedule(request, event_id):
    """Public schedule page"""
    event = get_object_or_404(Event, id=event_id)
    activities = event.activities.all().order_by('start_time')
    
    context = {
        'event': event,
        'activities': activities,
    }
    return render(request, 'website/event_schedule.html', context)


def event_website_gallery(request, event_id):
    """Public gallery page"""
    event = get_object_or_404(Event, id=event_id)
    
    context = {
        'event': event,
    }
    return render(request, 'website/event_gallery.html', context)


def event_website_contact(request, event_id):
    """Public contact page"""
    event = get_object_or_404(Event, id=event_id)
    
    context = {
        'event': event,
    }
    return render(request, 'website/event_contact.html', context)


def activity_register(request, activity_id):
    """Register for activity"""
    activity = get_object_or_404(Activity, id=activity_id)
    event = activity.event
    # Visitors are not allowed to register; redirect to login/participant signup
    if not request.user.is_authenticated or not _is_participant(request.user):
        messages.error(request, "Please log in as a Participant to register for activities.")
        return redirect(f"{reverse('events:unified_login')}?role=participant")

    existing_registration = ActivityRegistration.objects.filter(
        participant=request.user,
        activity=activity,
    ).first()
    payment_check = None
    if existing_registration:
        payment_check = getattr(existing_registration, "payment_check", None)

    if request.method == 'POST':
        name = (request.POST.get('name') or request.POST.get('full_name') or '').strip()
        email = (request.POST.get('email') or request.user.email or '').strip()
        phone = (request.POST.get('phone') or '').strip()
        college = (request.POST.get('college') or '').strip()
        payment_proof = request.FILES.get('payment_proof') or request.FILES.get('payment_qr')

        if not name or not email or not phone:
            messages.error(request, "Name, Email, and Phone are required.")
        elif activity.max_participants and ActivityRegistration.objects.filter(activity=activity).count() >= activity.max_participants and not existing_registration:
            messages.error(request, "Registration Closed: this activity is full.")
        else:
            registration = existing_registration
            if registration is None:
                registration = ActivityRegistration.objects.create(
                    participant=request.user,
                    activity=activity,
                    qr_token=uuid.uuid4(),
                )

            payment_check, _ = PaymentCheck.objects.get_or_create(registration=registration)
            if payment_proof:
                payment_check.payment_proof = payment_proof
                payment_check.save(update_fields=["payment_proof"])

            messages.success(request, "Registration submitted successfully.")
            existing_registration = registration
            return redirect("events:activity_register", activity_id=activity.id)

    context = {
        'activity': activity,
        'event': event,
        'registration_complete': bool(existing_registration),
        'download_pass_url': reverse("events:generate_activity_gate_pass", args=[existing_registration.id]) if existing_registration and existing_registration.status == ActivityRegistration.STATUS_APPROVED else None,
        'payment_check': payment_check,
    }
    return render(request, 'website/activity_register.html', context)


# ================================
# SIMPLE MULTI-STEP EVENT CREATION
# ================================

@login_required
def create_event_step1(request):
    if request.method == "POST":
        event = Event.objects.create(
            event=request.POST.get("event_name"),
            description=request.POST.get("description"),
            date_of_event=request.POST.get("date_of_event"),
            time_of_event=request.POST.get("time_of_event"),
            venue=request.POST.get("venue"),
            location=request.POST.get("location"),
            category=request.POST.get("category"),
            contact_info=request.POST.get("contact_info"),
            image_url=request.POST.get("image_url"),
            user=request.user,
        )

        request.session["event_id"] = event.id
        return redirect("events:create_event_step2")

    return render(request, "events/create_event.html")


@login_required
def create_event_step2(request):
    event_id = request.session.get("event_id")
    event = get_object_or_404(Event, id=event_id)

    if request.method == "POST":
        Activity.objects.create(
            event=event,
            name=request.POST.get("name"),
            description=request.POST.get("description"),
            date=request.POST.get("date"),
            start_time=request.POST.get("start_time"),
            end_time=request.POST.get("end_time"),
        )

        if request.POST.get("action") == "next":
            return redirect("events:create_event_step3")

    return render(request, "events/add_activity.html")


@login_required
def create_event_step3(request):
    event_id = request.session.get("event_id")

    if request.method == "POST":
        ActivityRegistrationFormField.objects.create(
            activity_id=event_id,
            label=request.POST.get("field_name"),
            field_type=request.POST.get("field_type"),
        )

        if request.POST.get("action") == "next":
            return redirect("events:create_event_step4")

    return render(request, "events/add_forms.html")


@login_required
def create_event_step4(request):
    event_id = request.session.get("event_id")

    if request.method == "POST":
        Coordinator.objects.create(
            event_id=event_id,
            name=request.POST.get("name"),
            email=request.POST.get("email"),
        )

        if request.POST.get("action") == "next":
            return redirect("events:create_event_step5")

    return render(request, "events/add_coordinator.html")


@login_required
def create_event_step5(request):
    event_id = request.session.get("event_id")
    event = get_object_or_404(Event, id=event_id)

    if request.method == "POST":
        event.template_choice = request.POST.get("template")
        event.save()

        return redirect("events:manage_events")

    return render(request, "events/website_setup.html")
